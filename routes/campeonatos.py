from flask import Blueprint, request, jsonify, make_response
from sqlalchemy.exc import OperationalError
from chaveamento import (
    alocar_partida_em_mesa,
    alocar_partida_grupo_em_mesa,
    avancar_para_mata_mata,
    gerar_chaveamento_vivo,
    gerar_fase_grupos,
    liberar_mesa_para_proxima_partida,
    liberar_mesa_partida_grupo,
    normalizar_categoria,
    obter_chaveamento_serializado,
    obter_estado_torneio,
)
from models import ChaveamentoPartida, PartidaGrupo, db, Campeonato, Mesa, Placar, JogadorInscrito, Categoria

bp = Blueprint('campeonatos', __name__, url_prefix='/api/campeonatos')


def normalizar_nivel(valor):
    nivel = (valor or 'iniciante').strip().lower()
    niveis_validos = {'iniciante', 'intermediario', 'avancado'}
    return nivel if nivel in niveis_validos else 'iniciante'


@bp.route('', methods=['GET'])
def listar_campeonatos():
    """Lista todos os campeonatos"""
    campeonatos = Campeonato.query.all()
    return jsonify([c.to_dict() for c in campeonatos])

@bp.route('', methods=['POST'])
def criar_campeonato():
    """Cria um novo campeonato"""
    dados = request.get_json()
    
    if not dados or 'nome' not in dados:
        return jsonify({'erro': 'Nome do campeonato é obrigatório'}), 400
    
    def _salvar_campeonato():
        novo = Campeonato(
            nome=dados['nome'],
            descricao=dados.get('descricao', '')
        )
        db.session.add(novo)
        db.session.commit()
        return novo

    try:
        novo_campeonato = _salvar_campeonato()
        return jsonify(novo_campeonato.to_dict()), 201
    except OperationalError:
        db.session.rollback()
        # Recupera schema em runtime e tenta novamente.
        db.create_all()
        novo_campeonato = _salvar_campeonato()
        return jsonify(novo_campeonato.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400

@bp.route('/<int:id>', methods=['GET'])
def obter_campeonato(id):
    """Obtém um campeonato específico"""
    campeonato = Campeonato.query.get(id)
    
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404
    
    return jsonify(campeonato.to_dict())

@bp.route('/<int:id>', methods=['PUT'])
def atualizar_campeonato(id):
    """Atualiza um campeonato"""
    campeonato = Campeonato.query.get(id)
    
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404
    
    dados = request.get_json()
    
    if 'nome' in dados:
        campeonato.nome = dados['nome']
    if 'descricao' in dados:
        campeonato.descricao = dados['descricao']
    if 'status' in dados:
        campeonato.status = dados['status']
    
    try:
        db.session.commit()
        return jsonify(campeonato.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400

@bp.route('/<int:id>', methods=['DELETE'])
def deletar_campeonato(id):
    """Deleta um campeonato e suas mesas"""
    campeonato = Campeonato.query.get(id)
    
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404
    
    try:
        db.session.delete(campeonato)
        db.session.commit()
        return jsonify({'mensagem': 'Campeonato deletado com sucesso'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400

@bp.route('/<int:id>/mesas', methods=['GET'])
def listar_mesas_campeonato(id):
    """Lista todas as mesas de um campeonato"""
    campeonato = Campeonato.query.get(id)
    
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404
    
    return jsonify([m.to_dict() for m in campeonato.mesas])

@bp.route('/<int:id>/jogadores-inscritos', methods=['GET'])
def listar_jogadores_inscritos(id):
    """Lista todos os jogadores inscritos em um campeonato"""
    campeonato = Campeonato.query.get(id)
    
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404
    
    jogadores = JogadorInscrito.query.filter_by(campeonato_id=id, ativo=True).all()
    return jsonify([j.to_dict() for j in jogadores])


@bp.route('/<int:id>/chaveamento', methods=['GET'])
def obter_chaveamento(id):
    """Obtém o chaveamento do campeonato, em modo simulado ou vivo."""
    campeonato = Campeonato.query.get(id)

    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404

    chaveamento = obter_chaveamento_serializado(id)
    chaveamento['campeonato_nome'] = campeonato.nome
    return jsonify(chaveamento)


@bp.route('/<int:id>/chaveamento-vivo', methods=['POST'])
def criar_chaveamento_vivo(id):
    """Gera ou regenera o chaveamento vivo do campeonato."""
    campeonato = Campeonato.query.get(id)

    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404

    dados = request.get_json(force=True, silent=True) or {}
    force = dados.get('force', False)

    if not force:
        partida_ativa = ChaveamentoPartida.query.filter(
            ChaveamentoPartida.campeonato_id == id,
            ChaveamentoPartida.status.in_(['em_andamento', 'finalizada'])
        ).first()
        if partida_ativa:
            return jsonify({
                'erro': 'Existem partidas em andamento ou finalizadas. Confirme para regenerar.',
                'requer_confirmacao': True
            }), 409

    try:
        chaveamento = gerar_chaveamento_vivo(id)
        db.session.commit()

        try:
            from app import broadcast_campeonato_update
            broadcast_campeonato_update(id, 'chaveamento_atualizado')
        except Exception as e:
            print(f"[BROADCAST ERROR] Erro ao notificar chaveamento: {e}")

        chaveamento['campeonato_nome'] = campeonato.nome
        return jsonify(chaveamento), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400


@bp.route('/<int:id>/chaveamento/partidas/<int:partida_id>/alocar-mesa', methods=['POST'])
def alocar_partida_chaveamento(id, partida_id):
    """Vincula uma partida do chaveamento a uma mesa e prepara os jogadores."""
    campeonato = Campeonato.query.get(id)

    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404

    dados = request.get_json() or {}
    mesa_id = dados.get('mesa_id')
    if not mesa_id:
        return jsonify({'erro': 'mesa_id é obrigatório'}), 400

    try:
        partida = alocar_partida_em_mesa(id, partida_id, int(mesa_id))
        db.session.commit()

        try:
            from app import broadcast_campeonato_update
            broadcast_campeonato_update(id, 'chaveamento_atualizado')
        except Exception as e:
            print(f"[BROADCAST ERROR] Erro ao notificar alocação de partida: {e}")

        return jsonify({
            'mensagem': 'Partida alocada à mesa com sucesso',
            'partida': partida.to_dict()
        })
    except ValueError as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400


@bp.route('/<int:id>/chaveamento/partidas/<int:partida_id>/liberar-mesa', methods=['POST'])
def liberar_mesa_partida_chaveamento(id, partida_id):
    """Desvincula a mesa de uma partida finalizada e a aloca na próxima partida pronta, se houver."""
    campeonato = Campeonato.query.get(id)

    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404

    try:
        proxima = liberar_mesa_para_proxima_partida(id, partida_id)
        db.session.commit()

        try:
            from app import broadcast_campeonato_update
            broadcast_campeonato_update(id, 'chaveamento_atualizado')
        except Exception as e:
            print(f"[BROADCAST ERROR] Erro ao notificar liberação de mesa: {e}")

        return jsonify({
            'mensagem': 'Mesa liberada com sucesso',
            'proxima_partida': proxima.to_dict() if proxima else None
        })
    except ValueError as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400


@bp.route('/<int:id>/torneio', methods=['GET'])
def obter_torneio(id):
    """Retorna o estado completo do torneio (fase de grupos + mata-mata)."""
    campeonato = Campeonato.query.get(id)
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404

    torneio = obter_estado_torneio(id)
    torneio['campeonato_nome'] = campeonato.nome
    return jsonify(torneio)


@bp.route('/<int:id>/fase-grupos', methods=['POST'])
def criar_fase_grupos(id):
    """Gera a fase de grupos para o campeonato."""
    campeonato = Campeonato.query.get(id)
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404

    dados = request.get_json(force=True, silent=True) or {}
    jogadores_por_grupo = int(dados.get('jogadores_por_grupo', 4))

    try:
        torneio = gerar_fase_grupos(id, jogadores_por_grupo=jogadores_por_grupo)
        db.session.commit()

        try:
            from app import broadcast_campeonato_update
            broadcast_campeonato_update(id, 'chaveamento_atualizado')
        except Exception as e:
            print(f"[BROADCAST ERROR] {e}")

        torneio['campeonato_nome'] = campeonato.nome
        return jsonify(torneio), 201
    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] Erro ao gerar fase de grupos para campeonato {id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 400


@bp.route('/<int:id>/fase-grupos', methods=['DELETE'])
def resetar_fase_grupos(id):
    """Deleta a fase de grupos do campeonato."""
    campeonato = Campeonato.query.get(id)
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404

    try:
        from sqlalchemy import text
        
        print(f"[RESET] Resetando fase de grupos para campeonato {id}")
        
        # Desabilitar foreign keys check para SQLite
        db.session.execute(text('PRAGMA foreign_keys = OFF'))
        
        # Deletar TODAS as classificações orphaned (cujo grupo não existe mais)
        orphaned_count = db.session.execute(
            text("DELETE FROM classificacoes_grupo WHERE grupo_id NOT IN (SELECT id FROM grupos_chaveamento)")
        )
        if orphaned_count.rowcount > 0:
            print(f"[RESET] Deletadas {orphaned_count.rowcount} classificação(ões) órfã(s)")
        
        # Deletar TODAS as classificações do campeonato
        result_class = db.session.execute(
            text("DELETE FROM classificacoes_grupo WHERE grupo_id IN (SELECT id FROM grupos_chaveamento WHERE campeonato_id = :cid)"),
            {'cid': id}
        )
        print(f"[RESET] Deletadas {result_class.rowcount} classificação(ões)")
        
        # Deletar partidas dos grupos
        result_partidas = db.session.execute(
            text("DELETE FROM partidas_grupo WHERE campeonato_id = :cid"),
            {'cid': id}
        )
        print(f"[RESET] Deletadas {result_partidas.rowcount} partida(s) de grupo")
        
        # Deletar todos os grupos do campeonato
        result_grupos = db.session.execute(
            text("DELETE FROM grupos_chaveamento WHERE campeonato_id = :cid"),
            {'cid': id}
        )
        print(f"[RESET] Deletados {result_grupos.rowcount} grupo(s)")
        
        # Última verificação e limpeza de órfãs
        final_orphaned = db.session.execute(
            text("DELETE FROM classificacoes_grupo WHERE grupo_id NOT IN (SELECT id FROM grupos_chaveamento)")
        )
        if final_orphaned.rowcount > 0:
            print(f"[RESET] Deletadas {final_orphaned.rowcount} classificação(ões) órfã(s) finais")
        
        # Reabilitar foreign keys check
        db.session.execute(text('PRAGMA foreign_keys = ON'))
        db.session.commit()

        try:
            from app import broadcast_campeonato_update
            broadcast_campeonato_update(id, 'chaveamento_atualizado')
        except Exception as e:
            print(f"[BROADCAST ERROR] {e}")

        print(f"[RESET] Fase de grupos resetada com sucesso")
        return jsonify({'mensagem': 'Fase de grupos resetada com sucesso'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] Erro ao resetar fase de grupos para campeonato {id}: {e}")
        import traceback
        traceback.print_exc()
        # Garantir que foreign keys seja reabilitada
        try:
            db.session.execute(text('PRAGMA foreign_keys = ON'))
        except:
            pass
        return jsonify({'erro': str(e)}), 400


@bp.route('/<int:id>/grupos/partidas/<int:partida_id>/alocar-mesa', methods=['POST'])
def alocar_partida_grupo(id, partida_id):
    """Aloca uma partida de grupo em uma mesa."""
    campeonato = Campeonato.query.get(id)
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404

    dados = request.get_json() or {}
    mesa_id = dados.get('mesa_id')
    if not mesa_id:
        return jsonify({'erro': 'mesa_id é obrigatório'}), 400

    try:
        partida = alocar_partida_grupo_em_mesa(id, partida_id, int(mesa_id))
        db.session.commit()

        try:
            from app import broadcast_campeonato_update
            broadcast_campeonato_update(id, 'chaveamento_atualizado')
        except Exception as e:
            print(f"[BROADCAST ERROR] {e}")

        return jsonify({'mensagem': 'Partida alocada com sucesso', 'partida': partida.to_dict()})
    except ValueError as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400


@bp.route('/<int:id>/grupos/partidas/<int:partida_id>/liberar-mesa', methods=['POST'])
def liberar_mesa_grupo(id, partida_id):
    """Libera a mesa de uma partida de grupo finalizada."""
    campeonato = Campeonato.query.get(id)
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404

    try:
        liberar_mesa_partida_grupo(id, partida_id)
        db.session.commit()

        try:
            from app import broadcast_campeonato_update
            broadcast_campeonato_update(id, 'chaveamento_atualizado')
        except Exception as e:
            print(f"[BROADCAST ERROR] {e}")

        return jsonify({'mensagem': 'Mesa liberada com sucesso'})
    except ValueError as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400


@bp.route('/<int:id>/grupos/partidas/<int:partida_id>/desalocar-mesa', methods=['POST'])
def desalocar_mesa_grupo(id, partida_id):
    """Desaloca (remove) a mesa de uma partida de grupo, sem precisar que esteja finalizada."""
    campeonato = Campeonato.query.get(id)
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404

    try:
        from models import PartidaGrupo, JogadorMesa
        from chaveamento import _resetar_placar_mesa
        
        partida = PartidaGrupo.query.get(partida_id)
        if not partida or partida.campeonato_id != id:
            return jsonify({'erro': 'Partida não encontrada'}), 404
        
        if not partida.mesa_id:
            return jsonify({'erro': 'Esta partida não está alocada a nenhuma mesa'}), 400
        
        mesa = partida.mesa
        
        # Remove os jogadores da mesa
        JogadorMesa.query.filter_by(mesa_id=mesa.id).delete(synchronize_session=False)
        
        # Reseta o placar
        _resetar_placar_mesa(mesa)
        
        # Marca a mesa como disponível
        mesa.status = 'disponivel'
        
        # Remove a alocação da partida (volta para 'pronta')
        partida.mesa_id = None
        partida.status = 'pronta'
        
        db.session.commit()

        try:
            from app import broadcast_campeonato_update
            broadcast_campeonato_update(id, 'chaveamento_atualizado')
        except Exception as e:
            print(f"[BROADCAST ERROR] {e}")

        return jsonify({
            'sucesso': True,
            'mensagem': f'Jogo desalocado da Mesa {mesa.numero}. Partida voltou para status pronta.',
            'partida': partida.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        print(f"[DESALOCAR ERROR] {str(e)}")
        return jsonify({'erro': str(e)}), 400


@bp.route('/<int:id>/avancar-mata-mata', methods=['POST'])
def avancar_mata_mata(id):
    """Gera o chaveamento eliminatório com os classificados dos grupos."""
    campeonato = Campeonato.query.get(id)
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404

    dados = request.get_json(force=True, silent=True) or {}
    qtd_avancam = int(dados.get('qtd_avancam', 2))

    try:
        torneio = avancar_para_mata_mata(id, qtd_avancam=qtd_avancam)
        db.session.commit()

        try:
            from app import broadcast_campeonato_update
            broadcast_campeonato_update(id, 'chaveamento_atualizado')
        except Exception as e:
            print(f"[BROADCAST ERROR] {e}")

        torneio['campeonato_nome'] = campeonato.nome
        return jsonify(torneio)
    except ValueError as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400


# ==================== ENDPOINTS DE CATEGORIAS ====================

@bp.route('/<int:campeonato_id>/categorias', methods=['GET'])
def listar_categorias(campeonato_id):
    """Lista todas as categorias de um campeonato"""
    campeonato = Campeonato.query.get(campeonato_id)
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404
    
    categorias = Categoria.query.filter_by(campeonato_id=campeonato_id).all()
    return jsonify([c.to_dict() for c in categorias])


@bp.route('/<int:campeonato_id>/categorias', methods=['POST'])
def criar_categoria(campeonato_id):
    """Cria uma nova categoria para um campeonato"""
    campeonato = Campeonato.query.get(campeonato_id)
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404
    
    dados = request.get_json()
    if not dados or 'nome' not in dados:
        return jsonify({'erro': 'Nome da categoria é obrigatório'}), 400
    
    try:
        categoria = Categoria(
            campeonato_id=campeonato_id,
            nome=dados['nome'],
            descricao=dados.get('descricao', '')
        )
        db.session.add(categoria)
        db.session.commit()
        return jsonify(categoria.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        if 'uq_campeonato_categoria_nome' in str(e):
            return jsonify({'erro': f'Categoria "{dados["nome"]}" já existe neste campeonato'}), 400
        return jsonify({'erro': str(e)}), 400


@bp.route('/categorias/<int:categoria_id>', methods=['GET'])
def obter_categoria(categoria_id):
    """Obtém uma categoria específica"""
    categoria = Categoria.query.get(categoria_id)
    if not categoria:
        return jsonify({'erro': 'Categoria não encontrada'}), 404
    
    return jsonify(categoria.to_dict())


@bp.route('/categorias/<int:categoria_id>', methods=['PUT'])
def atualizar_categoria(categoria_id):
    """Atualiza uma categoria"""
    categoria = Categoria.query.get(categoria_id)
    if not categoria:
        return jsonify({'erro': 'Categoria não encontrada'}), 404
    
    dados = request.get_json()
    if 'nome' in dados:
        categoria.nome = dados['nome']
    if 'descricao' in dados:
        categoria.descricao = dados['descricao']
    
    try:
        db.session.commit()
        return jsonify(categoria.to_dict())
    except Exception as e:
        db.session.rollback()
        if 'uq_campeonato_categoria_nome' in str(e):
            return jsonify({'erro': f'Categoria "{dados.get("nome")}" já existe neste campeonato'}), 400
        return jsonify({'erro': str(e)}), 400


@bp.route('/categorias/<int:categoria_id>', methods=['DELETE'])
def deletar_categoria(categoria_id):
    """Deleta uma categoria"""
    categoria = Categoria.query.get(categoria_id)
    if not categoria:
        return jsonify({'erro': 'Categoria não encontrada'}), 404
    
    # Verificar se há jogadores inscritos nesta categoria
    jogadores_count = JogadorInscrito.query.filter_by(categoria_id=categoria_id).count()
    if jogadores_count > 0:
        return jsonify({'erro': f'Não é possível deletar: há {jogadores_count} jogador(es) inscrito(s) nesta categoria'}), 400
    
    try:
        db.session.delete(categoria)
        db.session.commit()
        return jsonify({'mensagem': 'Categoria deletada com sucesso'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400

@bp.route('/<int:id>/jogadores-inscritos', methods=['POST'])
def adicionar_jogador_inscrito(id):
    """Adiciona um novo jogador ao campeonato"""
    campeonato = Campeonato.query.get(id)
    
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404
    
    dados = request.get_json()
    
    if not dados or 'nome' not in dados:
        return jsonify({'erro': 'Nome do jogador é obrigatório'}), 400
    
    try:
        novo_jogador = JogadorInscrito(
            nome=dados['nome'],
            categoria_id=dados.get('categoria_id'),
            categoria=dados.get('categoria'),  # Legado
            nivel=normalizar_nivel(dados.get('nivel')),
            campeonato_id=id,
            ativo=True
        )
        db.session.add(novo_jogador)
        db.session.commit()
        
        return jsonify(novo_jogador.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400

@bp.route('/<int:id>/jogadores-inscritos/<int:jogador_id>', methods=['DELETE'])
def remover_jogador_inscrito(id, jogador_id):
    """Remove um jogador inscrito do campeonato"""
    jogador = JogadorInscrito.query.get(jogador_id)
    
    if not jogador or jogador.campeonato_id != id:
        return jsonify({'erro': 'Jogador não encontrado'}), 404
    
    try:
        # Marcar como inativo ao invés de deletar
        jogador.ativo = False
        db.session.commit()
        
        return jsonify({'mensagem': 'Jogador removido com sucesso'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400

@bp.route('/<int:id>/jogadores-inscritos/<int:jogador_id>', methods=['PUT'])
def atualizar_jogador_inscrito(id, jogador_id):
    """Atualiza um jogador inscrito"""
    jogador = JogadorInscrito.query.get(jogador_id)
    
    if not jogador or jogador.campeonato_id != id:
        return jsonify({'erro': 'Jogador não encontrado'}), 404
    
    dados = request.get_json()
    
    if 'nome' in dados:
        jogador.nome = dados['nome']
    if 'categoria' in dados:
        jogador.categoria = normalizar_categoria(dados.get('categoria'))
    if 'nivel' in dados:
        jogador.nivel = normalizar_nivel(dados.get('nivel'))
    if 'ativo' in dados:
        jogador.ativo = dados['ativo']
    
    try:
        db.session.commit()
        return jsonify(jogador.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400

@bp.route('/<int:id>/importar-jogadores-excel', methods=['POST'])
def importar_jogadores_excel(id):
    """
    Importa jogadores inscritos a partir de um arquivo Excel.
    
    Esperado no arquivo Excel:
    - Coluna A: Nome do jogador
    - Coluna B: Nível (opcional: iniciante, intermediario, avancado)
    - Coluna C: Categoria (opcional: ID ou nome da categoria)
    
    O arquivo deve ser enviado como multipart/form-data com chave 'arquivo'
    """
    campeonato = Campeonato.query.get(id)
    
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404
    
    # Verificar se foi enviado um arquivo
    if 'arquivo' not in request.files:
        return jsonify({'erro': 'Nenhum arquivo foi enviado. Use a chave "arquivo"'}), 400
    
    arquivo = request.files['arquivo']
    
    if arquivo.filename == '':
        return jsonify({'erro': 'Arquivo não selecionado'}), 400
    
    # Verificar extensão do arquivo
    if not arquivo.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'erro': 'O arquivo deve ser no formato Excel (.xlsx ou .xls)'}), 400
    
    try:
        from openpyxl import load_workbook
        from io import BytesIO
        
        # Ler o arquivo Excel
        conteudo = arquivo.read()
        wb = load_workbook(BytesIO(conteudo))
        ws = wb.active
        
        jogadores_adicionados = []
        erros = []
        linhas_vazias = 0
        linhas_processadas = 0
        
        # Iterar sobre as linhas (começando da linha 2, pulando o cabeçalho)
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Verificar se a linha está completamente vazia
            if not row or not any(cell is not None for cell in row):
                linhas_vazias += 1
                continue
                
            if not row[0]:  # Se a coluna A (nome) está vazia, pula
                continue
            
            linhas_processadas += 1
            
            try:
                nome = str(row[0]).strip()
                
                if not nome:
                    erros.append(f"Linha {idx}: Nome vazio")
                    continue
                
                # Coluna B: Nível
                nivel_raw = row[1] if len(row) > 1 else None
                nivel = normalizar_nivel(nivel_raw)
                
                # Coluna C: Categoria
                categoria_id = None
                if len(row) > 2 and row[2]:
                    categoria_raw = row[2]
                    
                    # Se for número, trata como ID
                    if isinstance(categoria_raw, (int, float)):
                        categoria_id = int(categoria_raw)
                        # Validar se categoria existe neste campeonato
                        categoria = Categoria.query.filter_by(
                            id=categoria_id, 
                            campeonato_id=id
                        ).first()
                        if not categoria:
                            erros.append(f"Linha {idx}: Categoria com ID {categoria_id} não existe neste campeonato")
                            continue
                    else:
                        # Se for string, busca pelo nome
                        categoria_nome = str(categoria_raw).strip()
                        categoria = Categoria.query.filter_by(
                            nome=categoria_nome,
                            campeonato_id=id
                        ).first()
                        if categoria:
                            categoria_id = categoria.id
                        else:
                            # Se categoria não existe, criar um aviso mas continuar
                            erros.append(f"Linha {idx}: Categoria '{categoria_nome}' não encontrada - jogador será adicionado sem categoria")
                
                # Verificar se jogador já existe (pelo nome + categoria)
                jogador_existente = JogadorInscrito.query.filter_by(
                    nome=nome,
                    campeonato_id=id,
                    categoria_id=categoria_id
                ).first()
                
                if jogador_existente:
                    if jogador_existente.ativo:
                        erros.append(f"Linha {idx}: Jogador '{nome}' já está inscrito")
                    else:
                        # Reativar se estava inativo
                        jogador_existente.ativo = True
                        jogador_existente.nivel = nivel
                        db.session.add(jogador_existente)
                        jogadores_adicionados.append({
                            'nome': nome,
                            'nivel': nivel,
                            'categoria_id': categoria_id,
                            'status': 'reativado'
                        })
                else:
                    # Criar novo jogador
                    novo_jogador = JogadorInscrito(
                        nome=nome,
                        categoria_id=categoria_id,
                        nivel=nivel,
                        campeonato_id=id,
                        ativo=True
                    )
                    db.session.add(novo_jogador)
                    jogadores_adicionados.append({
                        'nome': nome,
                        'nivel': nivel,
                        'categoria_id': categoria_id,
                        'status': 'adicionado'
                    })
            
            except Exception as e:
                erros.append(f"Linha {idx}: Erro ao processar - {str(e)}")
        
        # Confirmar alterações no banco
        if jogadores_adicionados:
            db.session.commit()
        
        # Preparar mensagem informativa
        mensagem = ''
        if linhas_processadas == 0 and linhas_vazias > 0:
            mensagem = 'Arquivo contém apenas linhas vazias'
        elif len(jogadores_adicionados) > 0:
            mensagem = f'{len(jogadores_adicionados)} jogador(es) processado(s)'
        else:
            mensagem = 'Nenhum jogador foi adicionado. Verifique se o arquivo contém dados válidos.'
        
        resposta = {
            'mensagem': mensagem,
            'jogadores_adicionados': len([j for j in jogadores_adicionados if j['status'] == 'adicionado']),
            'jogadores_reativados': len([j for j in jogadores_adicionados if j['status'] == 'reativado']),
            'erros_count': len(erros),
            'jogadores': jogadores_adicionados,
            'linhas_processadas': linhas_processadas,
            'linhas_vazias': linhas_vazias
        }
        
        if erros:
            resposta['erros'] = erros
        
        status_code = 201 if len(jogadores_adicionados) > 0 else 400
        return jsonify(resposta), status_code
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'erro': f'Erro ao processar arquivo Excel: {str(e)}'
        }), 400

@bp.route('/<int:id>/template-jogadores', methods=['GET'])
def baixar_template_jogadores(id):
    """
    Retorna um arquivo Excel de template para importação de jogadores.
    O arquivo contém exemplos de como preencher os dados.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    campeonato = Campeonato.query.get(id)
    if not campeonato:
        return jsonify({'erro': 'Campeonato não encontrado'}), 404
    
    try:
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Jogadores'
        
        # Definir largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        
        # Estilos
        header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Cabeçalho
        headers = ['Nome', 'Nível', 'Categoria']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Exemplos de dados
        exemplos = [
            ['João Silva', 'iniciante', 'Sub-11'],
            ['Maria Santos', 'intermediario', 'Sub-13'],
            ['Pedro Costa', 'avancado', 'Adulto'],
            ['Ana Lima', 'iniciante', 'Sub-11'],
            ['Carlos Mendes', 'intermediario', 'Adulto'],
        ]
        
        # Preenchimento de exemplos
        for row_idx, exemplo in enumerate(exemplos, start=2):
            for col_idx, valor in enumerate(exemplo, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border
        
        # Adicionar linhas vazias para o usuário preencher
        for row_idx in range(len(exemplos) + 2, len(exemplos) + 7):
            for col_idx in range(1, 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
        
        # Adicionar abas de instrução
        ws_instrucoes = wb.create_sheet('Instruções')
        
        instrucoes = [
            ['INSTRUÇÕES PARA IMPORTAÇÃO DE JOGADORES'],
            [],
            ['Coluna A - NOME (Obrigatório)'],
            ['  • Nome completo do jogador'],
            ['  • Não pode estar vazio'],
            ['  • Máximo 100 caracteres'],
            [],
            ['Coluna B - NÍVEL (Opcional)'],
            ['  • Valores aceitos: iniciante, intermediario, avancado'],
            ['  • Padrão: iniciante (se deixar em branco)'],
            ['  • Não é sensível a maiúsculas'],
            [],
            ['Coluna C - CATEGORIA (Opcional)'],
            ['  • Pode ser o nome exato da categoria'],
            ['  • Ou o ID numérico da categoria'],
            ['  • Se não preenchido, sem categoria'],
            ['  • Deve existir no campeonato'],
        ]
        
        for row_idx, linha in enumerate(instrucoes, start=1):
            for col_idx, valor in enumerate(linha, start=1):
                cell = ws_instrucoes.cell(row=row_idx, column=col_idx, value=valor)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=12)
        
        ws_instrucoes.column_dimensions['A'].width = 50
        
        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Retornar arquivo com response manual
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename="modelo_jogadores_{campeonato.nome}.xlsx"'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Length'] = len(output.getvalue())
        return response
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao gerar template: {str(e)}'}), 400
