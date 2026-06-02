import pytest
import io
from openpyxl import Workbook
from app import app, db
from models import Campeonato, Categoria, JogadorInscrito


@pytest.fixture
def client():
    """Fixture para o cliente de teste da aplicação"""
    app.config['TESTING'] = True
    with app.app_context():
        db.create_all()
        yield app.test_client()
        db.session.remove()
        db.drop_all()


@pytest.fixture
def setup_data(client):
    """Fixture para criar dados de teste"""
    with app.app_context():
        # Criar um campeonato
        campeonato = Campeonato(nome='Campeonato Teste', descricao='Teste')
        db.session.add(campeonato)
        db.session.commit()
        
        # Criar categorias
        cat1 = Categoria(campeonato_id=campeonato.id, nome='Sub-11', descricao='Categoria Sub-11')
        cat2 = Categoria(campeonato_id=campeonato.id, nome='Adulto', descricao='Categoria Adulto')
        db.session.add(cat1)
        db.session.add(cat2)
        db.session.commit()
        
        return {
            'campeonato_id': campeonato.id,
            'categoria_1_id': cat1.id,
            'categoria_2_id': cat2.id
        }


def criar_arquivo_excel_teste():
    """Cria um arquivo Excel de teste em memória"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Jogadores'
    
    # Cabeçalho
    ws['A1'] = 'Nome'
    ws['B1'] = 'Nível'
    ws['C1'] = 'Categoria'
    
    # Dados
    ws['A2'] = 'João Silva'
    ws['B2'] = 'iniciante'
    ws['C2'] = 'Sub-11'
    
    ws['A3'] = 'Maria Santos'
    ws['B3'] = 'intermediario'
    ws['C3'] = 'Adulto'
    
    ws['A4'] = 'Pedro Costa'
    ws['B4'] = 'avancado'
    ws['C4'] = 'Adulto'
    
    # Salvar em BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def test_importar_jogadores_excel_sucesso(client, setup_data):
    """Test: Importação com sucesso de jogadores"""
    campeonato_id = setup_data['campeonato_id']
    
    arquivo = criar_arquivo_excel_teste()
    
    response = client.post(
        f'/api/campeonatos/{campeonato_id}/importar-jogadores-excel',
        data={'arquivo': (arquivo, 'teste.xlsx')},
        content_type='multipart/form-data'
    )
    
    assert response.status_code == 201
    data = response.get_json()
    
    # Verificações
    assert 'jogadores_adicionados' in data
    assert 'erros_count' in data
    assert data['jogadores_adicionados'] == 3
    assert data['erros_count'] == 0
    assert len(data['jogadores']) == 3
    
    # Verificar se foram salvos no banco
    with app.app_context():
        jogadores = JogadorInscrito.query.filter_by(campeonato_id=campeonato_id).all()
        assert len(jogadores) == 3
        nomes = [j.nome for j in jogadores]
        assert 'João Silva' in nomes
        assert 'Maria Santos' in nomes
        assert 'Pedro Costa' in nomes


def test_importar_jogadores_sem_arquivo(client, setup_data):
    """Test: Erro quando nenhum arquivo é enviado"""
    campeonato_id = setup_data['campeonato_id']
    
    response = client.post(
        f'/api/campeonatos/{campeonato_id}/importar-jogadores-excel',
        data={},
        content_type='multipart/form-data'
    )
    
    assert response.status_code == 400
    data = response.get_json()
    assert 'erro' in data
    assert 'Nenhum arquivo' in data['erro']


def test_importar_jogadores_arquivo_invalido(client, setup_data):
    """Test: Erro com arquivo inválido"""
    campeonato_id = setup_data['campeonato_id']
    
    arquivo = io.BytesIO(b'conteudo invalido')
    
    response = client.post(
        f'/api/campeonatos/{campeonato_id}/importar-jogadores-excel',
        data={'arquivo': (arquivo, 'teste.txt')},
        content_type='multipart/form-data'
    )
    
    assert response.status_code == 400
    data = response.get_json()
    assert 'erro' in data
    assert 'Excel' in data['erro'] or 'xlsx' in data['erro'].lower()


def test_importar_jogadores_campeonato_inexistente(client):
    """Test: Erro quando campeonato não existe"""
    arquivo = criar_arquivo_excel_teste()
    
    response = client.post(
        f'/api/campeonatos/999/importar-jogadores-excel',
        data={'arquivo': (arquivo, 'teste.xlsx')},
        content_type='multipart/form-data'
    )
    
    assert response.status_code == 404
    data = response.get_json()
    assert 'erro' in data
    assert 'não encontrado' in data['erro'].lower()


def test_importar_jogadores_duplicado(client, setup_data):
    """Test: Aviso quando jogador já existe"""
    campeonato_id = setup_data['campeonato_id']
    categoria_id = setup_data['categoria_1_id']
    
    # Adicionar jogador existente
    with app.app_context():
        jogador = JogadorInscrito(
            nome='João Silva',
            campeonato_id=campeonato_id,
            categoria_id=categoria_id,
            nivel='iniciante'
        )
        db.session.add(jogador)
        db.session.commit()
    
    arquivo = criar_arquivo_excel_teste()
    
    response = client.post(
        f'/api/campeonatos/{campeonato_id}/importar-jogadores-excel',
        data={'arquivo': (arquivo, 'teste.xlsx')},
        content_type='multipart/form-data'
    )
    
    data = response.get_json()
    
    # Deve ter um erro sobre jogador duplicado
    assert len(data['erros']) > 0
    assert any('já está inscrito' in erro.lower() for erro in data['erros'])


def test_importar_jogadores_categoria_invalida(client, setup_data):
    """Test: Erro quando categoria não existe"""
    campeonato_id = setup_data['campeonato_id']
    
    wb = Workbook()
    ws = wb.active
    
    ws['A1'] = 'Nome'
    ws['B1'] = 'Nível'
    ws['C1'] = 'Categoria'
    
    ws['A2'] = 'João Silva'
    ws['B2'] = 'iniciante'
    ws['C2'] = 'Categoria Inexistente'  # Categoria que não existe
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = client.post(
        f'/api/campeonatos/{campeonato_id}/importar-jogadores-excel',
        data={'arquivo': (output, 'teste.xlsx')},
        content_type='multipart/form-data'
    )
    
    data = response.get_json()
    
    # Deve ter um erro sobre categoria não encontrada
    assert len(data['erros']) > 0
    assert any('não encontrada' in erro.lower() for erro in data['erros'])


def test_baixar_template_jogadores(client, setup_data):
    """Test: Download de template Excel"""
    campeonato_id = setup_data['campeonato_id']
    
    response = client.get(f'/api/campeonatos/{campeonato_id}/template-jogadores')
    
    assert response.status_code == 200
    assert response.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert len(response.data) > 0
    
    # Verificar se é um arquivo Excel válido
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(response.data))
    assert 'Jogadores' in wb.sheetnames
    assert 'Instruções' in wb.sheetnames


def test_baixar_template_campeonato_inexistente(client):
    """Test: Erro ao baixar template de campeonato inexistente"""
    response = client.get('/api/campeonatos/999/template-jogadores')
    
    assert response.status_code == 404
    data = response.get_json()
    assert 'erro' in data


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
